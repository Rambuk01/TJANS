# import packages
import openpyxl
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import random
import time
from pygame import mixer
import csv

# import functions
#from excelhandler import openFileXl

DELTAGERE = []
DRAWN = []
NUM = 1
RADIO_TO_CELLS = {1:"C10", 2:"C12",3:"C13",4:"F2",5:"F4",6:"F5",7:"F6",8:"F8",9:"F9",10:"F10",11:"F12",12:"F13", 13:"I2", 14:"I4", 15:"I5", 16:"I6",17:"I8", 18:"I9", 19:"I10", 20:"I12", 21:"I13",22:"I14"}

mixer.init()
try:
    mixer.music.load('dice.ogg')
except:
    print("dice.ogg lydfil mangler")
def auto_fix(*args):
    #Denne funktion skal automatisk trække lod ml. elever og indsætte dem i Viggo
    print('Boop boop, automatisk lodtrækning og indsætning i VIGGO')


def openFileXl(*args):
    file_path = filedialog.askopenfilename()
    wb = load_workbook(filename=file_path)
    sheet_ranges = wb['ViGGO']
    for row in sheet_ranges.rows:
        if row[4].value == "Jeg er på NSE":
            deltager = "{} {}".format(row[0].value, row[1].value)
            print(deltager + " tilføjet")
            DELTAGERE.append(deltager)
    print("{} elever tilføjet".format(len(DELTAGERE)))
    #update_text()
    

def update_text():
    textfield.config(state=NORMAL)
    textfield.delete(1.0, END);
    counter = 1
    for deltager in DELTAGERE:
        textfield.insert(END, str(counter) + ". " + deltager + "\n");
        counter += 1
    textfield.config(state=DISABLED)

def tilføjElev(*args):
    if nytNavn.get():
        DELTAGERE.append(nytNavn.get())
        textfield.config(state=NORMAL)
        textfield.insert(END, nytNavn.get() + "\n");
        textfield.config(state=DISABLED)
        nytNavn.set("")

def clearDrawn():
    del DRAWN[:]
    drawfield.config(state=NORMAL)
    drawfield.delete(1.0, END);
    drawfield.config(state=DISABLED)

def drawstudents():
    nr = int(number_to_draw.get());
    drawfield.config(state=NORMAL)

    for number in range(nr):
        rnd = int(random.randrange(0, len(DELTAGERE)))
        student = DELTAGERE.pop(rnd)
        DRAWN.append(student)
        drawfield.insert(END, student + "\n")

    try:
        mixer.music.play()
    except:
        print("Kunne ikke afspille lydfil")

    time.sleep(1.75)
    drawfield.config(state=DISABLED)
    update_text()

def fjernElev(*e):
    print("Du vil gerne fjerne en elev.")
    try:
        elevnr = int(nytNavn.get())
        print(DELTAGERE.pop(elevnr) + " fjernet.")
        update_text()
    except:
        print("Cant turn it into int")

def insertXL(*e):
    """Lav de korte navne"""
    navne_til_celle = ""
    for deltager in DRAWN:
        navne_til_celle += ", {} {}".format(deltager.split()[0], deltager.split()[-1][0])
    navne_til_celle = navne_til_celle[1:]

    """Her skriver vi kode, hvor vi tager de trukkede elever og lægger dem
    i en bestemt celle i XL filen"""
    wb = openpyxl.load_workbook('TJANS.xlsx')
    sheet = wb['NSE']
    if tjans.get() != 0:
        """ Jeg skal indsætte navne_til_celle i den givne celle. """
        celle = RADIO_TO_CELLS[tjans.get()]
        sheet[celle].value = navne_til_celle
        try:
            wb.save('TJANS.xlsx')
            wb.close()
            print("{} indsat i {}".format(navne_til_celle, celle))
        except:
            print("Kunne ikke gemme...start forfra og sørg for lukke dokumentet TJANS.xlsx")
    else:
        print("Du har ikke angivet hvilken tjans eleven skal udføre")
    clearDrawn()


def callback(*args):
        root.quit()
        root.destroy()
        mixer.quit
        

def createWindow():
    # --- globals ---
    global root
    global filnavn
    global textfield
    global drawfield
    global nytNavn
    global number_to_draw
    global clear
    global fredag_radio_aftensmad_ud_ind
    global tjans
    global remove

    root = Tk()
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)

    # --- mainframe ---
    mainframe = ttk.Frame(root, padding=(3, 3, 12, 12))
    mainframe.grid(column=0, row=0, sticky=(N,W,E,S))

    # --- Widgest ---

    textfield = Text(mainframe, font=("Helvetica", 16), height=15, width=40);
    textfield.config(state=DISABLED)
    drawfield = Text(mainframe, font=("Helvetica", 24), height=10, width=28);
    drawfield.config(state=DISABLED)
    calc = ttk.Button(mainframe, text="Hent fil", command=openFileXl)
    tilføj = ttk.Button(mainframe, text="Tilføj", command=tilføjElev)
    draw = ttk.Button(mainframe, text="Træk", command=drawstudents)
    clear = ttk.Button(mainframe, text="Reset", command=clearDrawn)
    insert = ttk.Button(mainframe, text="Indsæt", command=insertXL)
    auto = ttk.Button(mainframe, text='AUTO', command=auto_fix)
    remove = ttk.Button(mainframe, text="Fjern", command=fjernElev)


    number_to_draw = Spinbox(mainframe, width=3, from_=1, to=10)
    info_label = ttk.Label(mainframe)
    radio_labelframe_left = LabelFrame(mainframe, bd=0, text="Tjanser: Fredag-Lørdag", padx=8, pady=5)
    radio_labelframe_right = LabelFrame(mainframe, bd=0, text="Tjanser: Søndag", padx=8, pady=5)
    nytNavn = StringVar()
    addName = ttk.Entry(mainframe, textvariable=nytNavn)

    # --- Radio Buttons ---
    tjans = IntVar()
    tjans.set(0)
    fredag_radio_aftensmad_ud_ind = Radiobutton(radio_labelframe_left, indicatoron=0, text="Fredag: Aftensmad MAD UD/IND", variable=tjans, value=1, width=28)
    fredag_radio_aftensmad_opvask = Radiobutton(radio_labelframe_left, indicatoron=0, text="Fredag: Aftensmad OPVASK", variable=tjans, value=2, width=28)
    fredag_radio_aft_bord_gulve = Radiobutton(radio_labelframe_left, indicatoron=0, text="Fredag: Aftensmad BORD/GULV", variable=tjans, value=3, width=28)
    lørdag_radio_mrgn_ud_ind = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Morgen MAD UD/IND", variable= tjans, value=4, width=28)
    lørdag_radio_mrgn_opvask = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Morgen OPVASK", variable=tjans, value=5, width=28)
    lørdag_radio_mrgn_bord_gulve = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Morgen BORD/GULV", variable=tjans, value=6, width=28)
    lørdag_radio_mdg_ud_ind = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Middag MAD UD/IND", variable=tjans, value=7, width=28)
    lørdag_radio_mdg_opvask = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Middag OPVASK", variable=tjans, value=8, width=28)
    lørdag_radio_mdg_bord_gulve = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Middag BORD/GULV", variable=tjans, value=9, width=28)
    lørdag_radio_aft_ud_ind = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Aftensmad MAD UD/IND", variable=tjans, value=10, width=28)
    lørdag_radio_aft_opvask = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Aftensmad OPVASK", variable=tjans, value=11, width=28)
    lørdag_radio_aft_bord_gulve = Radiobutton(radio_labelframe_left, indicatoron=0, text="Lørdag: Aftensmad BORD/GULV", variable=tjans, value=12, width=28)
    søndag_radio_mrgn_ud_ind = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Morgen MAD UD/IND", variable=tjans, value=13, width=28)
    søndag_radio_mrgn_opvask = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Morgen OPVASK", variable=tjans, value=14, width=28)
    søndag_radio_mrgn_bord_gulve = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Morgen BORD/GULV", variable=tjans, value=15, width=28)
    søndag_radio_mdg_ud_ind = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Middag MAD UD/IND", variable=tjans, value=16, width=28)
    søndag_radio_mdg_opvask = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Middag OPVASK", variable=tjans, value=17, width=28)
    søndag_radio_mdg_bord_gulve = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Middag BORD/GULV", variable=tjans, value=18, width=28)
    søndag_radio_aft_ud_ind = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Aftensmad MAD UD/IND", variable=tjans, value=19, width=28)
    søndag_radio_aft_opvask = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Aftensmad OPVASK", variable=tjans, value=20, width=28)
    søndag_radio_aft_bord_gulve = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: Aftensmad BORD/GULV", variable=tjans, value=21, width=28)
    søndag_radio_nat_oprydning = Radiobutton(radio_labelframe_right, indicatoron=0, text="Søndag: 22 OPRYDNING", variable=tjans, value=22, width=28)
    
    # --- Setup grid ---
    textfield.grid(column=0, row=0, columnspan=4)
    drawfield.grid(column=4, row=0, columnspan=4)
    calc.grid(column=0, row=1, sticky=W)
    tilføj.grid(column=0, row=2, sticky=W)
    addName.grid(column=1, row=2, sticky=(W,E))
    remove.grid(column=2, row=2, sticky=(W,E))
    draw.grid(column=4, row=1, sticky=W)
    clear.grid(column=4, row=1, sticky=E)
    number_to_draw.grid(column=4, row=2, sticky=W)
    insert.grid(column=5, row=1, sticky=E)
    auto.grid(column=5, row=2, sticky=E)

    # ---Radiobutton grid ---
    radio_labelframe_left.grid(column=9, row=0, rowspan=2, sticky=(W,N))
    radio_labelframe_right.grid(column=10, row=0, rowspan=2, sticky=(W,N))
    fredag_radio_aftensmad_ud_ind.grid(column=0, row=0, sticky=(W,N))
    fredag_radio_aftensmad_opvask.grid(column=0, row=1, sticky=(W,N))
    fredag_radio_aft_bord_gulve.grid(column=0, row=2, sticky=(W,N))
    lørdag_radio_mrgn_ud_ind.grid(column=0, row=3, sticky=(W,N))
    lørdag_radio_mrgn_opvask.grid(column=0, row=4, sticky=(W,N))
    lørdag_radio_mrgn_bord_gulve.grid(column=0, row=5, sticky=(W,N))
    lørdag_radio_mdg_ud_ind.grid(column=0, row=6, sticky=(W,N))
    lørdag_radio_mdg_opvask.grid(column=0, row=7, sticky=(W,N))
    lørdag_radio_mdg_bord_gulve.grid(column=0, row=8, sticky=(W,N))
    lørdag_radio_aft_ud_ind.grid(column=0, row=9, sticky=(W,N))
    lørdag_radio_aft_opvask.grid(column=0, row=10, sticky=(W,N))
    lørdag_radio_aft_bord_gulve.grid(column=0, row=11, sticky=(W,N))
    søndag_radio_mrgn_ud_ind.grid(column=0, row=1, sticky=(W,N))
    søndag_radio_mrgn_opvask.grid(column=0, row=2, sticky=(W,N))
    søndag_radio_mrgn_bord_gulve.grid(column=0, row=3, sticky=(W,N))
    søndag_radio_mdg_ud_ind.grid(column=0, row=4, sticky=(W,N))
    søndag_radio_mdg_opvask.grid(column=0, row=5, sticky=(W,N))
    søndag_radio_mdg_bord_gulve.grid(column=0, row=6, sticky=(W,N))
    søndag_radio_aft_ud_ind.grid(column=0, row=7, sticky=(W,N))
    søndag_radio_aft_opvask.grid(column=0, row=8, sticky=(W,N))
    søndag_radio_aft_bord_gulve.grid(column=0, row=9, sticky=(W,N))
    søndag_radio_nat_oprydning.grid(column=0, row=10, sticky=(W,N))
    # --- Binds ---
    root.bind("<Return>", tilføjElev)

    # --- Loop ---
    root.bind("<Escape>", callback)
    root.bind("<Return>", insertXL)
    root.protocol("WM_DELETE_WINDOW", callback)
    root.mainloop()


createWindow()
